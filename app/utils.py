# ===========================================
# utils.py - Các hàm bổ trợ
# ===========================================
# File này chứa các utility functions như:
# - Xuất Excel
# - Định dạng thời gian
# - Các helper functions khác

import os
import io
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from flask import url_for


def format_datetime(dt, format_str='%d/%m/%Y %H:%M:%S'):
    """
    Định dạng datetime thành chuỗi
    
    Args:
        dt: Đối tượng datetime
        format_str: Chuỗi định dạng (mặc định: dd/mm/yyyy HH:MM:SS)
        
    Returns:
        str: Chuỗi thời gian đã định dạng
    """
    if dt is None:
        return "N/A"
    
    if isinstance(dt, str):
        return dt
    
    try:
        return dt.strftime(format_str)
    except:
        return str(dt)


def format_datetime_vi(dt):
    """
    Định dạng datetime theo kiểu Việt Nam
    
    Args:
        dt: Đối tượng datetime
        
    Returns:
        str: Chuỗi thời gian tiếng Việt (vd: "30/01/2026 14:30:00")
    """
    return format_datetime(dt, '%d/%m/%Y %H:%M:%S')


def format_relative_time(dt, lang='vi'):
    """
    Chuyển đổi datetime thành thời gian tương đối
    
    Args:
        dt: Đối tượng datetime
        lang: Ngôn ngữ ('vi', 'en', 'zh', 'ja')
        
    Returns:
        str: Chuỗi thời gian tương đối (vd: "5 phút trước", "2 giờ trước")
    """
    if dt is None:
        return "N/A"
    
    now = datetime.utcnow()
    diff = now - dt
    
    seconds = diff.total_seconds()
    
    if lang == 'en':
        if seconds < 60:
            return "Just now"
        elif seconds < 3600:
            minutes = int(seconds / 60)
            return f"{minutes} minute{'s' if minutes > 1 else ''} ago"
        elif seconds < 86400:
            hours = int(seconds / 3600)
            return f"{hours} hour{'s' if hours > 1 else ''} ago"
        elif seconds < 604800:
            days = int(seconds / 86400)
            return f"{days} day{'s' if days > 1 else ''} ago"
        elif seconds < 2592000:
            weeks = int(seconds / 604800)
            return f"{weeks} week{'s' if weeks > 1 else ''} ago"
        else:
            return format_datetime_vi(dt)
    elif lang == 'zh':
        if seconds < 60:
            return "刚刚"
        elif seconds < 3600:
            minutes = int(seconds / 60)
            return f"{minutes}分钟前"
        elif seconds < 86400:
            hours = int(seconds / 3600)
            return f"{hours}小时前"
        elif seconds < 604800:
            days = int(seconds / 86400)
            return f"{days}天前"
        elif seconds < 2592000:
            weeks = int(seconds / 604800)
            return f"{weeks}周前"
        else:
            return format_datetime_vi(dt)
    elif lang == 'ja':
        if seconds < 60:
            return "たった今"
        elif seconds < 3600:
            minutes = int(seconds / 60)
            return f"{minutes}分前"
        elif seconds < 86400:
            hours = int(seconds / 3600)
            return f"{hours}時間前"
        elif seconds < 604800:
            days = int(seconds / 86400)
            return f"{days}日前"
        elif seconds < 2592000:
            weeks = int(seconds / 604800)
            return f"{weeks}週間前"
        else:
            return format_datetime_vi(dt)
    else:
        if seconds < 60:
            return "Vừa xong"
        elif seconds < 3600:
            minutes = int(seconds / 60)
            return f"{minutes} phút trước"
        elif seconds < 86400:
            hours = int(seconds / 3600)
            return f"{hours} giờ trước"
        elif seconds < 604800:
            days = int(seconds / 86400)
            return f"{days} ngày trước"
        elif seconds < 2592000:
            weeks = int(seconds / 604800)
            return f"{weeks} tuần trước"
        else:
            return format_datetime_vi(dt)


def export_to_excel(comments, scrape_info=None):
    """
    Xuất danh sách comment ra file Excel chuyên nghiệp
    
    Args:
        comments: List các dictionary chứa thông tin comment
        scrape_info: Dictionary chứa thông tin về lần scrape (optional)
        
    Returns:
        io.BytesIO: Buffer chứa file Excel
    """
    # Tạo workbook mới
    wb = Workbook()
    ws = wb.active
    ws.title = "Comments"
    
    # Định nghĩa styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Thêm thông tin scrape ở đầu file (nếu có)
    current_row = 1
    if scrape_info:
        info_font = Font(bold=True, size=11)
        
        ws.cell(row=current_row, column=1, value="Thông tin Scrape")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
        current_row += 1
        
        info_items = [
            ("Platform:", scrape_info.get('platform', 'N/A').upper()),
            ("URL:", scrape_info.get('url', 'N/A')),
            ("Thời gian scrape:", format_datetime_vi(scrape_info.get('scraped_at', datetime.utcnow()))),
            ("Tổng số comment:", str(len(comments))),
        ]
        
        for label, value in info_items:
            ws.cell(row=current_row, column=1, value=label).font = info_font
            ws.cell(row=current_row, column=2, value=value)
            current_row += 1
        
        current_row += 1  # Dòng trống
    
    # Headers cho bảng comment (bỏ cột Thời gian)
    headers = ["STT", "Username", "Nội dung", "Lượt thích"]
    header_row = current_row
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Đổ dữ liệu comment
    for idx, comment in enumerate(comments, 1):
        row = header_row + idx
        
        # STT
        ws.cell(row=row, column=1, value=idx).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=1).border = thin_border
        
        # Username
        ws.cell(row=row, column=2, value=comment.get('username', 'N/A'))
        ws.cell(row=row, column=2).alignment = cell_alignment
        ws.cell(row=row, column=2).border = thin_border
        
        # Nội dung
        ws.cell(row=row, column=3, value=comment.get('content', ''))
        ws.cell(row=row, column=3).alignment = cell_alignment
        ws.cell(row=row, column=3).border = thin_border
        
        # Lượt thích
        ws.cell(row=row, column=4, value=comment.get('likes', 0))
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4).border = thin_border
    
    # Điều chỉnh độ rộng cột
    ws.column_dimensions['A'].width = 8   # STT
    ws.column_dimensions['B'].width = 25  # Username
    ws.column_dimensions['C'].width = 60  # Nội dung
    ws.column_dimensions['D'].width = 12  # Lượt thích
    
    # Lưu vào buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


def export_to_excel_simple(comments):
    """
    Xuất comment ra Excel đơn giản bằng pandas
    
    Args:
        comments: List các dictionary chứa thông tin comment
        
    Returns:
        io.BytesIO: Buffer chứa file Excel
    """
    # Tạo DataFrame
    df = pd.DataFrame(comments)
    
    # Đổi tên cột sang tiếng Việt
    column_mapping = {
        'username': 'Tên người dùng',
        'content': 'Nội dung',
        'timestamp': 'Thời gian',
        'likes': 'Lượt thích'
    }
    df = df.rename(columns=column_mapping)
    
    # Thêm cột STT
    df.insert(0, 'STT', range(1, len(df) + 1))
    
    # Xuất ra buffer
    buffer = io.BytesIO()
    df.to_excel(buffer, index=False, engine='openpyxl')
    buffer.seek(0)
    
    return buffer


def validate_url(url):
    """
    Kiểm tra URL có hợp lệ không
    
    Args:
        url: URL cần kiểm tra
        
    Returns:
        tuple: (is_valid: bool, platform: str hoặc None, error_message: str hoặc None)
    """
    if not url:
        return False, None, "URL không được để trống"
    
    url = url.strip()
    
    # Kiểm tra có phải URL không
    if not url.startswith(('http://', 'https://')):
        return False, None, "URL phải bắt đầu bằng http:// hoặc https://"
    
    url_lower = url.lower()
    
    # Kiểm tra platform
    if 'tiktok.com' in url_lower:
        return True, 'tiktok', None
    elif 'facebook.com' in url_lower or 'fb.watch' in url_lower:
        return True, 'facebook', None
    else:
        return False, None, "URL phải là link từ TikTok hoặc Facebook"


def sanitize_filename(filename):
    """
    Làm sạch tên file, loại bỏ các ký tự không hợp lệ
    
    Args:
        filename: Tên file gốc
        
    Returns:
        str: Tên file đã được làm sạch
    """
    # Các ký tự không được phép trong tên file
    invalid_chars = '<>:"/\\|?*'
    
    for char in invalid_chars:
        filename = filename.replace(char, '_')
    
    # Giới hạn độ dài
    if len(filename) > 200:
        filename = filename[:200]
    
    return filename


def generate_export_filename(platform, url=None):
    """
    Tạo tên file export tự động
    
    Args:
        platform: Tên platform (tiktok/facebook)
        url: URL gốc (optional)
        
    Returns:
        str: Tên file export
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    return f"comments_{platform}_{timestamp}.xlsx"


def truncate_text(text, max_length=100):
    """
    Cắt ngắn text nếu quá dài
    
    Args:
        text: Text gốc
        max_length: Độ dài tối đa
        
    Returns:
        str: Text đã được cắt ngắn
    """
    if not text:
        return ""
    
    if len(text) <= max_length:
        return text
    
    return text[:max_length - 3] + "..."


def get_status_badge_class(status):
    """
    Lấy class CSS cho status badge
    
    Args:
        status: Trạng thái (pending/success/failed)
        
    Returns:
        str: Class CSS tương ứng
    """
    status_classes = {
        'pending': 'bg-yellow-100 text-yellow-800',
        'success': 'bg-green-100 text-green-800',
        'failed': 'bg-red-100 text-red-800'
    }
    
    return status_classes.get(status, 'bg-gray-100 text-gray-800')


def get_platform_icon(platform):
    """
    Lấy icon cho platform
    
    Args:
        platform: Tên platform
        
    Returns:
        str: URL path đến icon
    """
    icons = {
        'tiktok': url_for('static', filename='icon/icon-tiktok.png'),
        'facebook': url_for('static', filename='icon/icon-facebook.png')
    }
    
    return icons.get(platform.lower(), url_for('static', filename='icon/icon-tiktok.png'))
